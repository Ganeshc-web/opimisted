import io

import pytest
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool

from app import db
from app.config import Config, config_map
from app.middleware.auth import hash_key
from app.models.api_key import APIKey
from app.models.assessment import AssessmentRecord
from app.models.communication import CommunicationDetails
from app.models.rate_config import RateConfig
from app.services.assessment_import_service import import_assessments_from_upload
from app.services.excel_service import rows_to_xlsx_bytes, spreadsheet_to_pdf_bytes


class ExcelTestConfig(Config):
    TESTING = True
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_ENGINE_OPTIONS = {
        "connect_args": {"check_same_thread": False},
        "poolclass": StaticPool,
    }


ADMIN_KEY = "excel-test-admin-key"


@pytest.fixture()
def app():
    config_map["excel_test"] = ExcelTestConfig
    from run import create_app

    test_app = create_app("excel_test")
    with test_app.app_context():
        db.create_all()
        db.session.add(
            APIKey(
                client_name="Admin",
                key_hash=hash_key(ADMIN_KEY),
                role="admin",
                is_active=True,
            )
        )
        db.session.add(
            RateConfig(
                inflation_post=0.06,
                roi_post=0.08,
                inflation_pre=0.06,
                roi_pre=0.12,
                pf_growth=0.05,
            )
        )
        db.session.commit()
        yield test_app
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def client(app):
    return app.test_client()


@pytest.fixture()
def admin_headers():
    return {"X-API-Key": ADMIN_KEY}


def test_rates_include_pf_growth(client, admin_headers):
    response = client.get("/api/v1/rates/", headers=admin_headers)
    body = response.get_json()
    assert response.status_code == 200
    assert body["data"]["pf_growth"] == 0.05


def test_rates_update_pf_growth(client, admin_headers):
    response = client.put(
        "/api/v1/rates/",
        json={"pf_growth": 0.07},
        headers=admin_headers,
    )
    body = response.get_json()
    assert response.status_code == 200
    assert body["data"]["pf_growth"] == 0.07


def test_rows_to_xlsx_bytes():
    payload = rows_to_xlsx_bytes(
        [{"name": "Test User", "email": "test@example.com"}],
        sheet_name="Users",
    )
    wb = load_workbook(io.BytesIO(payload))
    assert wb.sheetnames == ["Users"]
    assert wb["Users"]["A2"].value == "Test User"


def test_spreadsheet_csv_to_pdf():
    class FakeUpload:
        filename = "sample.csv"

        def read(self):
            return b"name,email\nAlice,alice@example.com\n"

    pdf_bytes, name = spreadsheet_to_pdf_bytes(FakeUpload())
    assert name == "sample.pdf"
    assert pdf_bytes.startswith(b"%PDF")


def test_import_assessments_from_csv(app):
    class FakeUpload:
        filename = "clients.csv"

        def read(self):
            return (
                b"mobile,email,consent\n"
                b"9876543210,import1@example.com,true\n"
                b"9123456780,import2@example.com,false\n"
            )

    with app.app_context():
        result = import_assessments_from_upload(FakeUpload())
        assert result["created"] == 2
        assert len(result["assessment_ids"]) == 2
        assert CommunicationDetails.query.count() == 2
        assert AssessmentRecord.query.count() == 2


def test_admin_import_assessments_endpoint(client, admin_headers):
    csv_bytes = (
        b"mobile,email,consent\n"
        b"9876543210,admin-import@example.com,yes\n"
    )
    response = client.post(
        "/api/v1/admin/upload/import-assessments",
        data={"file": (io.BytesIO(csv_bytes), "clients.csv")},
        headers=admin_headers,
        content_type="multipart/form-data",
    )
    body = response.get_json()
    assert response.status_code == 200
    assert body["data"]["created"] == 1
    assert len(body["data"]["assessment_ids"]) == 1


def test_admin_import_template_download(client, admin_headers):
    response = client.get(
        "/api/v1/admin/upload/import-template",
        headers=admin_headers,
    )
    assert response.status_code == 200
    assert (
        response.headers["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(response.data))
    assert wb.sheetnames == ["Clients"]
    assert wb["Clients"]["A1"].value == "mobile"


def test_admin_import_missing_columns(client, admin_headers):
    response = client.post(
        "/api/v1/admin/upload/import-assessments",
        data={"file": (io.BytesIO(b"name,email\nAlice,a@x.com\n"), "bad.csv")},
        headers=admin_headers,
        content_type="multipart/form-data",
    )
    body = response.get_json()
    assert response.status_code == 400
    assert body["code"] == "INVALID_INPUT"


def test_admin_convert_pdf_route_imports_assessments(client, admin_headers):
    """Legacy upload URL used by admin frontend should import, not return PDF."""
    csv_bytes = (
        b"mobile,email,consent\n"
        b"9876501234,legacy-import@example.com,true\n"
    )
    response = client.post(
        "/api/v1/admin/upload/convert-pdf",
        data={"file": (io.BytesIO(csv_bytes), "clients.csv")},
        headers=admin_headers,
        content_type="multipart/form-data",
    )
    body = response.get_json()
    assert response.status_code == 200
    assert body["data"]["created"] == 1
    assert len(body["data"]["assessment_ids"]) == 1
