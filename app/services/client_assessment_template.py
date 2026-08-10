"""Client-facing assessment Excel template (admin → customer → upload).

Designed for a non-technical customer: pre-labelled goals, no category/type/
inflation/child-number columns. Backend maps labels to assessment fields.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Matches product goal catalogue; inflation is applied on import (not shown to client).
GOAL_TEMPLATE_ROWS: list[dict[str, Any]] = [
    # Children's education & marriage (inflation 8% education / 6% marriage)
    {
        "label": "Child 1's Graduation",
        "category": "child_goal",
        "goal_type": "Graduation",
        "child_number": 1,
        "inflation_rate": 0.08,
        "section": "Children's Education and Marriage Goals",
    },
    {
        "label": "Child 1's Post Graduation",
        "category": "child_goal",
        "goal_type": "Post Graduation",
        "child_number": 1,
        "inflation_rate": 0.08,
        "section": "Children's Education and Marriage Goals",
    },
    {
        "label": "Child 1's Marriage",
        "category": "child_goal",
        "goal_type": "Marriage",
        "child_number": 1,
        "inflation_rate": 0.06,
        "section": "Children's Education and Marriage Goals",
    },
    {
        "label": "Child 2's Graduation",
        "category": "child_goal",
        "goal_type": "Graduation",
        "child_number": 2,
        "inflation_rate": 0.08,
        "section": "Children's Education and Marriage Goals",
    },
    {
        "label": "Child 2's Post Graduation",
        "category": "child_goal",
        "goal_type": "Post Graduation",
        "child_number": 2,
        "inflation_rate": 0.08,
        "section": "Children's Education and Marriage Goals",
    },
    {
        "label": "Child 2's Marriage",
        "category": "child_goal",
        "goal_type": "Marriage",
        "child_number": 2,
        "inflation_rate": 0.06,
        "section": "Children's Education and Marriage Goals",
    },
    # Other / lifestyle goals (inflation 6%)
    {
        "label": "Home Purchase",
        "category": "lifestyle",
        "goal_type": "Home Purchase",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Car Purchase",
        "category": "lifestyle",
        "goal_type": "Car Purchase",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Home Renovation",
        "category": "lifestyle",
        "goal_type": "Home Renovation",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Holiday Home",
        "category": "lifestyle",
        "goal_type": "Holiday Home",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Foreign Tour",
        "category": "lifestyle",
        "goal_type": "Foreign Tour",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Family Gifting",
        "category": "lifestyle",
        "goal_type": "Family Gifting",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Charity",
        "category": "lifestyle",
        "goal_type": "Charity",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Child Birth Expenses",
        "category": "lifestyle",
        "goal_type": "Child Birth Expenses",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Big Purchases",
        "category": "lifestyle",
        "goal_type": "Big Purchases",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Estate For Children",
        "category": "lifestyle",
        "goal_type": "Estate For Children",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
    },
    {
        "label": "Others",
        "category": "lifestyle",
        "goal_type": "Other",
        "child_number": None,
        "inflation_rate": 0.06,
        "section": "Other Goals",
        "needs_custom_name": True,
    },
]

# Friendly label → internal field key (Your Details sheet).
CLIENT_DETAIL_FIELDS: list[tuple[str, str, str]] = [
    # (section, label shown to client, internal key)
    ("Contact", "Mobile number", "mobile"),
    ("Contact", "Email", "email"),
    ("Contact", "Consent to email report (Yes/No)", "consent"),
    ("Contact", "Spouse mobile", "spouse_mobile"),
    ("Contact", "Spouse email", "spouse_email"),
    ("Contact", "Residential address", "residential_address"),
    ("About you", "Full name", "client_name"),
    ("About you", "Occupation", "client_occupation"),
    ("About you", "Designation", "client_designation"),
    ("About you", "Company", "client_company"),
    ("About you", "Date of birth (DD/MM/YYYY)", "client_dob"),
    ("About you", "Retirement age", "client_retirement_age"),
    ("About your spouse", "Spouse full name", "spouse_name"),
    ("About your spouse", "Spouse occupation", "spouse_occupation"),
    ("About your spouse", "Spouse designation", "spouse_designation"),
    ("About your spouse", "Spouse company", "spouse_company"),
    ("About your spouse", "Spouse date of birth (DD/MM/YYYY)", "spouse_dob"),
    ("About your spouse", "Spouse retirement age", "spouse_retirement_age"),
    (
        "Retirement planning",
        "Annual retirement income needed — you (Rs)",
        "client_annual_ret_reqd",
    ),
    (
        "Retirement planning",
        "Annual retirement income needed — spouse (Rs)",
        "spouse_annual_ret_reqd",
    ),
    ("Retirement planning", "Monthly household expenses (Rs)", "household_monthly"),
    ("Retirement savings", "Your EPF yearly contribution (Rs)", "client_epf_annual"),
    ("Retirement savings", "Your current EPF balance (Rs)", "client_epf_accum"),
    ("Retirement savings", "Spouse EPF yearly contribution (Rs)", "spouse_epf_annual"),
    ("Retirement savings", "Spouse current EPF balance (Rs)", "spouse_epf_accum"),
    ("Retirement savings", "Employer NPS monthly (Rs)", "employer_nps_pm"),
    ("Retirement savings", "Your NPS monthly (Rs)", "self_nps_pm"),
    ("Retirement savings", "Current NPS balance (Rs)", "current_nps_accum"),
    ("Retirement savings", "Superannuation monthly (Rs)", "sa_pm"),
    ("Retirement savings", "Current Superannuation balance (Rs)", "current_sa_accum"),
    (
        "Retirement savings",
        "Spouse employer NPS monthly (Rs)",
        "spouse_employer_nps_pm",
    ),
    ("Retirement savings", "Spouse NPS monthly (Rs)", "spouse_self_nps_pm"),
    (
        "Retirement savings",
        "Spouse current NPS balance (Rs)",
        "spouse_current_nps_accum",
    ),
    ("Retirement savings", "Spouse Superannuation monthly (Rs)", "spouse_sa_pm"),
    (
        "Retirement savings",
        "Spouse current Superannuation balance (Rs)",
        "spouse_current_sa_accum",
    ),
]

# Sample answers so the downloaded file is self-explanatory.
SAMPLE_ANSWERS: dict[str, Any] = {
    "mobile": "9876543210",
    "email": "client@example.com",
    "consent": "Yes",
    "spouse_mobile": "9876543211",
    "spouse_email": "spouse@example.com",
    "residential_address": "123 Main St, Mumbai",
    "client_name": "Rahul Sharma",
    "client_occupation": "Engineer",
    "client_designation": "Manager",
    "client_company": "Tech Corp",
    "client_dob": "01/01/1990",
    "client_retirement_age": 60,
    "spouse_name": "Priya Sharma",
    "spouse_occupation": "Teacher",
    "spouse_designation": "Senior Teacher",
    "spouse_company": "ABC School",
    "spouse_dob": "01/01/1995",
    "spouse_retirement_age": 55,
    "client_annual_ret_reqd": 1500000,
    "spouse_annual_ret_reqd": 1000000,
    "household_monthly": 30000,
    "client_epf_annual": 33600,
    "client_epf_accum": 500000,
    "spouse_epf_annual": 7200,
    "spouse_epf_accum": 0,
    "employer_nps_pm": 5000,
    "self_nps_pm": 2000,
    "current_nps_accum": 150000,
    "sa_pm": 3000,
    "current_sa_accum": 200000,
}

HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
SECTION_FILL = PatternFill("solid", fgColor="FFE699")
TITLE_FILL = PatternFill("solid", fgColor="F8CBAD")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin", color="9E9E9E"),
    right=Side(style="thin", color="9E9E9E"),
    top=Side(style="thin", color="9E9E9E"),
    bottom=Side(style="thin", color="9E9E9E"),
)
BOLD = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
NORMAL = Font(name="Calibri", size=11)


def _style_cell(cell, *, fill=None, bold=False, center=False) -> None:
    cell.font = BOLD if bold else NORMAL
    cell.border = THIN
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    if fill is not None:
        cell.fill = fill


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _normalize_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("—", "-").replace("–", "-")
    return " ".join(text.split())


def goal_catalog_by_label() -> dict[str, dict[str, Any]]:
    """Map normalized goal label → metadata for import."""
    catalog: dict[str, dict[str, Any]] = {}
    aliases = {
        "child 1's pg": "Child 1's Post Graduation",
        "child 2's pg": "Child 2's Post Graduation",
        "house purchase": "Home Purchase",
        "other": "Others",
        "others": "Others",
    }
    for row in GOAL_TEMPLATE_ROWS:
        catalog[_normalize_label(row["label"])] = row
    for alias, canonical in aliases.items():
        for row in GOAL_TEMPLATE_ROWS:
            if row["label"] == canonical:
                catalog[alias] = row
                break
    return catalog


def client_field_label_map() -> dict[str, str]:
    """normalized label → internal key."""
    mapping: dict[str, str] = {}
    for _section, label, key in CLIENT_DETAIL_FIELDS:
        mapping[_normalize_label(label)] = key
        mapping[_normalize_label(key.replace("_", " "))] = key
    # Extra friendly aliases from older sheets / product screenshots.
    extras = {
        "mobile": "mobile",
        "phone": "mobile",
        "mobile number": "mobile",
        "email address": "email",
        "consent": "consent",
        "consent to email report": "consent",
        "full name": "client_name",
        "client name": "client_name",
        "date of birth": "client_dob",
        "dob": "client_dob",
        "annual retirement required": "client_annual_ret_reqd",
        "annual retirement income needed": "client_annual_ret_reqd",
        "annual retirement income needed - you": "client_annual_ret_reqd",
        "annual retirement income needed - spouse": "spouse_annual_ret_reqd",
        "spouse annual retirement required": "spouse_annual_ret_reqd",
        "monthly household expenses": "household_monthly",
        "household monthly": "household_monthly",
        "your epf yearly contribution": "client_epf_annual",
        "your current epf balance": "client_epf_accum",
        "employer nps monthly": "employer_nps_pm",
        "your nps monthly": "self_nps_pm",
        "current nps balance": "current_nps_accum",
        "superannuation monthly": "sa_pm",
        "current superannuation balance": "current_sa_accum",
    }
    mapping.update(extras)
    return mapping


def build_client_assessment_template_bytes(*, with_sample: bool = True) -> bytes:
    """Pretty 3-sheet workbook for customers to fill."""
    wb = Workbook()

    # --- Your Details ---
    ws = wb.active
    ws.title = "Your Details"
    ws["A1"] = "Wealth Wisdom — Client Assessment Form"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:C1")
    ws["A2"] = (
        "Fill the yellow cells only. Leave blank anything that does not apply. "
        "Dates must be DD/MM/YYYY. Then send this file back — we generate your report from it."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 40

    headers = ["Section", "Field", "Your answer"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(3, col, title)
        _style_cell(cell, fill=HEADER_FILL, bold=True, center=True)

    row_idx = 4
    last_section = None
    for section, label, key in CLIENT_DETAIL_FIELDS:
        if section != last_section:
            # Visual section break
            last_section = section
        ws.cell(row_idx, 1, section)
        ws.cell(row_idx, 2, label)
        answer = SAMPLE_ANSWERS.get(key, "") if with_sample else ""
        ws.cell(row_idx, 3, answer)
        for col in range(1, 4):
            fill = SECTION_FILL if col == 1 else (INPUT_FILL if col == 3 else None)
            _style_cell(ws.cell(row_idx, col), fill=fill, bold=(col == 1))
        row_idx += 1

    _set_widths(ws, {"A": 22, "B": 48, "C": 28})

    # --- Children ---
    ws_c = wb.create_sheet("Children")
    ws_c["A1"] = "Children"
    ws_c["A1"].font = TITLE_FONT
    ws_c["A1"].fill = TITLE_FILL
    ws_c.merge_cells("A1:E1")
    ws_c["A2"] = (
        "Fill one row per child. Child number is the row label (Child 1, Child 2…) — "
        "you do not need to enter a number. Leave unused rows blank."
    )
    ws_c["A2"].alignment = Alignment(wrap_text=True)
    ws_c.merge_cells("A2:E2")
    ws_c.row_dimensions[2].height = 36

    child_headers = [
        "Child",
        "Full name",
        "Date of birth (DD/MM/YYYY)",
        "Occupation",
        "Financially dependent (Yes/No)",
    ]
    for col, title in enumerate(child_headers, start=1):
        cell = ws_c.cell(3, col, title)
        _style_cell(cell, fill=HEADER_FILL, bold=True, center=True)

    sample_children = [
        ("Child 1", "Aarav Sharma", "01/06/2010", "Student", "Yes"),
        ("Child 2", "", "", "", "Yes"),
        ("Child 3", "", "", "", "Yes"),
    ]
    if not with_sample:
        sample_children = [
            ("Child 1", "", "", "", "Yes"),
            ("Child 2", "", "", "", "Yes"),
            ("Child 3", "", "", "", "Yes"),
        ]
    for r, values in enumerate(sample_children, start=4):
        for c, value in enumerate(values, start=1):
            cell = ws_c.cell(r, c, value)
            fill = SECTION_FILL if c == 1 else INPUT_FILL
            _style_cell(cell, fill=fill, bold=(c == 1))

    _set_widths(ws_c, {"A": 12, "B": 22, "C": 28, "D": 16, "E": 30})

    # --- Goals ---
    ws_g = wb.create_sheet("Goals")
    ws_g["A1"] = "Needs and Goals"
    ws_g["A1"].font = TITLE_FONT
    ws_g["A1"].fill = TITLE_FILL
    ws_g.merge_cells("A1:D1")
    ws_g["A2"] = (
        "Enter Target year and Current cost (today's Rs) only for goals you care about. "
        "Leave cost blank to skip a goal. Inflation and goal type are applied automatically. "
        "For “Others”, also write a short custom name."
    )
    ws_g["A2"].alignment = Alignment(wrap_text=True)
    ws_g.merge_cells("A2:D2")
    ws_g.row_dimensions[2].height = 48

    goal_headers = [
        "Goal",
        "Target year",
        "Current cost (Rs)",
        "Custom name (Others only)",
    ]
    for col, title in enumerate(goal_headers, start=1):
        cell = ws_g.cell(3, col, title)
        _style_cell(cell, fill=HEADER_FILL, bold=True, center=True)

    row_idx = 4
    last_section = None
    sample_goal_costs = {
        "Child 1's Graduation": (2035, 2500000),
        "Home Purchase": (2030, 5000000),
        "Others": (2031, 500000),
    }
    for meta in GOAL_TEMPLATE_ROWS:
        if meta["section"] != last_section:
            last_section = meta["section"]
            ws_g.cell(row_idx, 1, last_section)
            ws_g.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            for col in range(1, 5):
                _style_cell(ws_g.cell(row_idx, col), fill=SECTION_FILL, bold=True)
            row_idx += 1

        year, cost, custom = "", "", ""
        if with_sample and meta["label"] in sample_goal_costs:
            year, cost = sample_goal_costs[meta["label"]]
            if meta.get("needs_custom_name"):
                custom = "World Cup Trip"
        ws_g.cell(row_idx, 1, meta["label"])
        ws_g.cell(row_idx, 2, year)
        ws_g.cell(row_idx, 3, cost)
        ws_g.cell(row_idx, 4, custom)
        _style_cell(ws_g.cell(row_idx, 1), fill=None, bold=True)
        for col in (2, 3, 4):
            _style_cell(ws_g.cell(row_idx, col), fill=INPUT_FILL)
        row_idx += 1

    _set_widths(ws_g, {"A": 36, "B": 14, "C": 20, "D": 28})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
